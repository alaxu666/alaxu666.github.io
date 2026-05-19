import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl.utils import get_column_letter
import subprocess
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import glob


def extract_chinese(text):
    """提取字符串中的中文字符"""
    if pd.isna(text):
        return text
    return ''.join(re.findall(r'[\u4e00-\u9fff]+', str(text)))


def extract_six_digits(text):
    """从文本中提取6位连续数字，数字前后都是非数字"""
    if pd.isna(text):
        return ""
    
    text_str = str(text)
    # 使用正则表达式查找6位连续数字，且前后都是非数字或边界
    match = re.search(r'(?<!\d)(\d{6})(?!\d)', text_str)
    
    if match:
        return match.group(1)
    else:
        return ""


def extract_project_name_from_commas(text):
    """
    从Project Name中提取第二个","和连续6位数字前的","之间的文本
    格式示例：XXX, YYY, ZZZ, 123456, AAA → 提取"ZZZ"
    """
    if pd.isna(text):
        return ""
    
    text_str = str(text)
    
    # 找到所有逗号的位置
    comma_positions = [i for i, char in enumerate(text_str) if char == ',']
    
    # 需要至少3个逗号（第二个逗号和6位数字前的逗号）
    if len(comma_positions) < 3:
        # 如果格式不符合，返回原文本
        return text_str
    
    # 第二个逗号的位置（索引从0开始）
    second_comma = comma_positions[1]
    
    # 查找连续6位数字
    six_digit_match = re.search(r'(?<!\d)(\d{6})(?!\d)', text_str)
    if not six_digit_match:
        # 如果没有找到6位数字，返回第二个逗号之后的所有内容
        return text_str[second_comma + 1:].strip()
    
    six_digit_start = six_digit_match.start()
    
    # 查找6位数字前的最后一个逗号
    comma_before_six_digit = -1
    for i in range(len(comma_positions) - 1, -1, -1):
        if comma_positions[i] < six_digit_start:
            comma_before_six_digit = comma_positions[i]
            break
    
    if comma_before_six_digit == -1 or comma_before_six_digit <= second_comma:
        # 如果找不到合适的逗号，返回第二个逗号之后的内容
        return text_str[second_comma + 1:six_digit_start].strip()
    
    # 提取文本（去掉首尾空格）
    extracted = text_str[second_comma + 1:comma_before_six_digit].strip()
    return extracted


def get_ebp_leader_from_web(driver, project_id):
    """
    在Program Dashboard页面中搜索项目ID，获取EBP Leader信息
    """
    try:
        print(f"正在处理项目 {project_id}...")
        
        # 确保在正确的iframe中
        # 先切回默认内容，然后重新进入iframe嵌套
        driver.switch_to.default_content()
        
        # 重新进入iframeContent
        iframe_content = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "iframeContent"))
        )
        driver.switch_to.frame(iframe_content)
        
        # 确保XSO Management菜单已展开，Program Dashboard已选中
        # 如果不在Program Dashboard页面，重新点击
        try:
            # 检查是否在Program Dashboard页面（通过检查搜索框是否存在）
            driver.switch_to.default_content()
            iframe_content = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "iframeContent"))
            )
            driver.switch_to.frame(iframe_content)
            
            # 尝试切换到嵌套iframe
            try:
                nested_frame = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "iframe.iframe_tab_menu-1-1"))
                )
                driver.switch_to.frame(nested_frame)
                # 检查搜索框是否存在
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.ID, "searchProject"))
                )
            except:
                # 不在正确的页面，需要重新导航
                print(f"项目 {project_id}: 不在Program Dashboard页面，重新导航...")
                driver.switch_to.default_content()
                iframe_content = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "iframeContent"))
                )
                driver.switch_to.frame(iframe_content)
                
                # 点击XSO Management（如果还没展开）
                try:
                    xso_management = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[text()='XSO Management']/parent::a"))
                    )
                    xso_management.click()
                    time.sleep(1)
                except:
                    pass  # 可能已经展开了
                
                # 点击Program Dashboard
                program_dashboard = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//span[text()='Program Dashboard']/parent::a"))
                )
                program_dashboard.click()
                time.sleep(3)
                
                # 进入嵌套iframe
                nested_frame = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "iframe.iframe_tab_menu-1-1"))
                )
                driver.switch_to.frame(nested_frame)
        except Exception as e:
            print(f"项目 {project_id}: 导航到Program Dashboard失败: {str(e)}")
            return None
        
        # 现在应该在正确的iframe中，进行搜索
        # 定位搜索框
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "searchProject"))
        )
        search_input.clear()
        search_input.send_keys(str(project_id))
        
        # 点击搜索按钮
        search_button = driver.find_element(By.CSS_SELECTOR, "button[onclick='searchproject()']")
        search_button.click()
        
        # 等待搜索结果
        time.sleep(2)
        
        # 查找EBP Leader的td元素
        try:
            td_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "td[aria-describedby='grid-table_real_pm_name']"))
            )
            title_value = td_element.get_attribute("title")
            print(f"项目 {project_id}: 获取到EBP Leader: {title_value}")
            
            # 提取中文字符
            chinese_name = extract_chinese(title_value)
            return chinese_name if chinese_name else title_value
            
        except TimeoutException:
            print(f"项目 {project_id}: 未找到EBP Leader元素")
            
            # 尝试使用XPath作为备选
            try:
                td_element = driver.find_element(By.XPATH, "//td[@aria-describedby='grid-table_real_pm_name']")
                title_value = td_element.get_attribute("title")
                chinese_name = extract_chinese(title_value)
                return chinese_name if chinese_name else title_value
            except:
                print(f"项目 {project_id}: XPath方式也未找到元素")
                return None
                
    except Exception as e:
        print(f"项目 {project_id}: 处理过程中出现错误: {str(e)}")
        return None


def sync_with_feishu_project_table(df_prt):
    """
    同步project.xlsx数据到飞书项目信息表V5.xlsx
    按照要求：
    1. 对比Project ID的前8位数字
    2. 如果找到相同，复制指定的字段数据（但EBP Leader字段如果非空则保留原值）
    3. 把飞书表中的Project ID复制到"项目名"字段
    4. 从飞书表的Project Name中提取文本到"项目名"字段
    5. 如果没找到，新增一行
    6. 对于EBP Leader为空的记录，从网页获取并更新
    """
    # 飞书表格路径
    feishu_path = r'C:\XSR\githubPage\yanfeng\project 飞书项目信息表V5.xlsx'
    
    # 如果飞书表格不存在，创建一个空的
    if not os.path.exists(feishu_path):
        # 创建带必要列的DataFrame
        feishu_columns = [
            'Project ID', 'Project Name', 'EBP Leader', 'Product Engineer',
            'Project State', 'Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR',
            'Phase 4 Gate Exit-CPA', '工作令号', '项目名', '项目号'
        ]
        df_feishu = pd.DataFrame(columns=feishu_columns)
        df_feishu.to_excel(feishu_path, index=False, engine='openpyxl')
        print(f"创建新的飞书项目信息表: {feishu_path}")
    
    # 读取飞书表格
    try:
        df_feishu = pd.read_excel(feishu_path, engine='openpyxl')
        print(f"成功读取飞书项目信息表，共{len(df_feishu)}行数据")
    except Exception as e:
        print(f"读取飞书项目信息表失败: {str(e)}")
        return
    
    # 确保飞书表格有必要的列
    required_columns = [
        'Project ID', 'Project Name', 'EBP Leader', 'Product Engineer',
        'Project State', 'Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR',
        'Phase 4 Gate Exit-CPA', '工作令号', '项目名', '项目号'
    ]
    
    for col in required_columns:
        if col not in df_feishu.columns:
            df_feishu[col] = pd.NA
    
    # 准备要复制的字段列表（不包括Project ID）
    # 注意：EBP Leader被单独处理，不包含在批量复制中
    copy_fields = [
        'Project Name', 'Product Engineer', 'Project State',
        'Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR', 'Phase 4 Gate Exit-CPA', '工作令号'
    ]
    
    # 处理project.xlsx中的每个项目
    updated_count = 0
    added_count = 0
    
    for _, project_row in df_prt.iterrows():
        # 提取Project ID的前8位数字
        project_id_full = str(project_row['Project ID'])
        match = re.search(r'^\d{8}', project_id_full)
        
        if not match:
            print(f"跳过Project ID格式不正确: {project_id_full}")
            continue
        
        project_id_8digits = match.group(0)
        
        # 在飞书表格中查找匹配的Project ID（前8位）
        found = False
        for idx, feishu_row in df_feishu.iterrows():
            feishu_id = str(feishu_row['Project ID'])
            if feishu_id.startswith(project_id_8digits):
                # 找到匹配项，更新数据
                found = True
                
                # 1. 复制指定字段数据（不包括EBP Leader）
                for field in copy_fields:
                    if field in project_row:
                        try:
                            df_feishu.at[idx, field] = project_row[field]
                        except TypeError:
                            # 如果发生类型错误，先将该列转为 object 类型，再重试
                            df_feishu[field] = df_feishu[field].astype(object)
                            df_feishu.at[idx, field] = project_row[field]
                
                # 2. 单独处理EBP Leader字段：只有当飞书表格中为空时才更新
                current_ebp = df_feishu.at[idx, 'EBP Leader']
                is_ebp_empty = pd.isna(current_ebp) or str(current_ebp).strip() == ''
                
                if is_ebp_empty and 'EBP Leader' in project_row:
                    # 如果飞书中的EBP Leader为空，则从project_row更新
                    df_feishu.at[idx, 'EBP Leader'] = project_row['EBP Leader']
                    print(f"项目 {project_id_8digits}: EBP Leader为空，更新为 '{project_row['EBP Leader']}'")
                elif not is_ebp_empty:
                    # 如果飞书中的EBP Leader非空，保留原值
                    print(f"项目 {project_id_8digits}: EBP Leader非空，保留原值 '{current_ebp}'")
                
                # 3. 把飞书表中的Project ID复制到"项目名"字段
                df_feishu.at[idx, '项目名'] = str(feishu_row['Project ID'])
                
                # 4. 从飞书表的Project Name中提取文本到"项目名"字段
                project_name = str(feishu_row['Project Name'])
                extracted_name = extract_project_name_from_commas(project_name)
                df_feishu.at[idx, '项目名'] = extracted_name
                
                updated_count += 1
                print(f"更新项目: {project_id_8digits} (索引: {idx})")
                break
        
        if not found:
            # 没有找到匹配项，新增一行
            # 处理Project ID：只保留左边8位数字
            new_project_id = project_id_8digits if len(project_id_8digits) >= 8 else str(project_row['Project ID'])[:8]
            
            new_row = {
                'Project ID': new_project_id,  # 只保留左边8位数字
                'Project Name': project_row.get('Project Name', ''),
                'EBP Leader': project_row.get('EBP Leader', ''),
                'Product Engineer': project_row.get('Product Engineer', ''),
                'Project State': project_row.get('Project State', ''),
                'Phase 2 Gate Exit-DVR': project_row.get('Phase 2 Gate Exit-DVR', ''),
                'Phase 3 Gate Exit-FPR': project_row.get('Phase 3 Gate Exit-FPR', ''),
                'Phase 4 Gate Exit-CPA': project_row.get('Phase 4 Gate Exit-CPA', ''),
                '工作令号': project_row.get('工作令号', ''),
                '项目名': new_project_id,  # 初始填充处理后的Project ID
                '项目号': new_project_id   # 将Project ID数据复制到"项目号"列
            }
            
            # 将新行添加到DataFrame
            new_row_df = pd.DataFrame([new_row])
            df_feishu = pd.concat([df_feishu, new_row_df], ignore_index=True)
            
            added_count += 1
            print(f"新增项目: {new_project_id} (原始ID: {project_row['Project ID']})")
            
            # 获取新行的索引
            last_idx = len(df_feishu) - 1
            
            # 对于新行，Project Name来自project.xlsx，所以从project.xlsx的Project Name中提取
            project_name = str(project_row.get('Project Name', ''))
            extracted_name = extract_project_name_from_commas(project_name)
            
            # 更新"项目名"字段：先用处理后的Project ID，然后用提取的文本覆盖
            df_feishu.at[last_idx, '项目名'] = new_project_id
            if extracted_name and extracted_name != project_name:
                df_feishu.at[last_idx, '项目名'] = extracted_name
    
    print(f"飞书表格更新完成：更新{updated_count}行，新增{added_count}行")
    
    # ========== 新增：从网页获取EBP Leader ==========
    print("\n开始从网页获取EBP Leader信息...")
    
    # 找出EBP Leader为空的记录
    empty_ebp_mask = df_feishu['EBP Leader'].isna() | (df_feishu['EBP Leader'] == '') | (df_feishu['EBP Leader'].astype(str).str.strip() == '')
    empty_ebp_projects = df_feishu[empty_ebp_mask]
    
    if len(empty_ebp_projects) == 0:
        print("没有EBP Leader为空的记录，跳过网页获取步骤")
    else:
        print(f"找到{len(empty_ebp_projects)}条EBP Leader为空的记录")
        
        # 初始化浏览器
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--allow-insecure-localhost')
        
        driver = webdriver.Chrome(options=chrome_options)
        driver.maximize_window()
        
        try:
            # 1. 打开登录页面
            print("正在打开登录页面...")
            driver.get("https://plmcnprdawc.yanfeng.com:3000/#/showHome")
            time.sleep(3)
            
            # 2. 输入用户名
            print("正在输入用户名...")
            username_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='username']"))
            )
            username_input.clear()
            username_input.send_keys("uxuxs004")
            
            # 3. 输入密码
            print("正在输入密码...")
            password_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='Password']")
            password_input.clear()
            password_input.send_keys("ABCabc%123")
            
            # 4. 点击登录按钮
            print("正在点击登录按钮...")
            login_button = driver.find_element(By.CSS_SELECTOR, "button.sw-button.accent-caution")
            login_button.click()
            
            # 5. 验证登录成功
            print("正在验证登录...")
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[icon-id='homeEdit']"))
            )
            print("登录成功！")
            
            # 6. 点击title='XSO & PKR'的div
            print("正在点击XSO & PKR...")
            xso_div = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@title='XSO & PKR']"))
            )
            xso_div.click()
            time.sleep(3)
            
            # 7. 进入iframeContent
            print("正在切换到iframeContent...")
            iframe_content = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "iframeContent"))
            )
            driver.switch_to.frame(iframe_content)
            
            # 8. 点击XSO Management
            print("正在点击XSO Management...")
            xso_management = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='XSO Management']/parent::a"))
            )
            xso_management.click()
            time.sleep(2)
            
            # 9. 点击Program Dashboard
            print("正在点击Program Dashboard...")
            program_dashboard = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Program Dashboard']/parent::a"))
            )
            program_dashboard.click()
            time.sleep(3)
            
            # 10. 进入嵌套iframe
            print("正在切换到嵌套iframe...")
            nested_frame = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "iframe.iframe_tab_menu-1-1"))
            )
            driver.switch_to.frame(nested_frame)
            
            # 11. 循环处理每个EBP Leader为空的记录
            ebp_updated_count = 0
            for idx, row in empty_ebp_projects.iterrows():
                project_id = row['Project ID']
                print(f"\n处理第{ebp_updated_count + 1}/{len(empty_ebp_projects)}个项目: {project_id}")
                
                # 获取EBP Leader
                ebp_leader = get_ebp_leader_from_web(driver, project_id)
                
                if ebp_leader:
                    # 更新飞书表格
                    df_feishu.at[idx, 'EBP Leader'] = ebp_leader
                    ebp_updated_count += 1
                    print(f"项目 {project_id}: EBP Leader更新为 '{ebp_leader}'")
                else:
                    print(f"项目 {project_id}: 未能获取到EBP Leader")
                
                # 短暂暂停，避免操作太快
                time.sleep(1)
            
            print(f"\nEBP Leader更新完成：成功更新{ebp_updated_count}条记录")
            
        except Exception as e:
            print(f"网页操作过程中出现错误: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            driver.quit()
            print("浏览器已关闭")
    
    # ========== 保存飞书表格 ==========
    try:
        with pd.ExcelWriter(feishu_path, engine='openpyxl') as writer:
            df_feishu.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # 设置列宽
            column_widths = {
                'Project ID': 15,
                'Project Name': 40,
                'EBP Leader': 10,
                'Product Engineer': 10,
                'Project State': 12,
                'Phase 2 Gate Exit-DVR': 12,
                'Phase 3 Gate Exit-FPR': 12,
                'Phase 4 Gate Exit-CPA': 12,
                '工作令号': 10,
                '项目名': 30,
                '项目号': 15
            }
            
            for col_name, width in column_widths.items():
                if col_name in df_feishu.columns:
                    col_idx = df_feishu.columns.get_loc(col_name)
                    col_letter = get_column_letter(col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = width
            
            # 设置日期格式
            date_cols = ['Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR', 'Phase 4 Gate Exit-CPA']
            for col in date_cols:
                if col in df_feishu.columns:
                    col_idx = df_feishu.columns.get_loc(col)
                    col_letter = get_column_letter(col_idx + 1)
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = 'YYYY-MM-DD'
        
        print(f"飞书项目信息表保存成功")
        
    except Exception as e:
        print(f"保存飞书项目信息表失败: {str(e)}")


def download_project_list():
    """下载Project_List文件（完整浏览器操作流程）"""
    download_dir = r"C:\Users\uxuxs004\Downloads"
    os.makedirs(download_dir, exist_ok=True)

    # 1. 浏览器配置（改为Chrome）
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "safebrowsing.disable_download_protection": True
    })

    # 添加以下参数来禁用不安全下载的警告
    chrome_options.add_argument("--safebrowsing-disable-download-protection")
    chrome_options.add_argument("--safebrowsing-disable-extension-blacklist")
    chrome_options.add_argument("--disable-features=DownloadBubble,DownloadBubbleV2")

    # 如果上述方法仍然无效，可以尝试更强制的方法（不推荐在生产环境使用）
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--disable-web-security")

    # 设置自动确认下载（危险操作，但可以避免弹出窗口）
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,  # 禁用安全浏览
        "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_settings.popups": 0
    })

    # 2. 启动浏览器（改为Chrome）
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    driver.implicitly_wait(10)  # 隐式等待10秒

    # 添加JavaScript执行来设置下载行为
    driver.execute_cdp_cmd("Page.setDownloadBehavior", {
        "behavior": "allow",
        "downloadPath": download_dir
    })

    try:
        # 3. 打开登录页面
        print("正在打开登录页面...")
        driver.get("https://plmcnprdawc.yanfeng.com:3000/#/showHome")

        # 4. 输入用户名
        print("正在输入用户名...")
        username = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='username']"))
        )
        username.clear()
        username.send_keys("uxuxs004")  # 替换为您的实际用户名

        # 5. 输入密码
        print("正在输入密码...")
        password = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder='Password']"))
        )
        password.clear()
        password.send_keys("ABCabc%123")  # 替换为您的实际密码

        # 6. 点击登录按钮
        print("正在登录...")
        login_btn = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.sw-button.accent-caution"))
        )
        login_btn.click()

        # 7. 验证登录成功
        print("验证登录...")
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "[icon-id='homeEdit']"))
        )
        print("登录成功")

        # 8. 导航到XSO & PKR页面
        print("正在跳转到XSO & PKR页面...")
        xso_link = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "[title='XSO & PKR']"))
        )
        driver.execute_script("arguments[0].click();", xso_link)

        # 9. 等待iframe加载并切换
        print("处理iframe...")
        WebDriverWait(driver, 15).until(
            EC.frame_to_be_available_and_switch_to_it((By.ID, "iframeContent"))
        )

        # 10. 展开XSO Management菜单
        print("展开菜单...")
        menu = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='XSO Management']/parent::a"))
        )
        menu.click()

        # 11. 点击Program Dashboard
        print("选择Program Dashboard...")
        dashboard = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='Program Dashboard']/parent::a"))
        )
        dashboard.click()
        time.sleep(3)  # 等待页面加载

        # 12. 处理可能的嵌套iframe
        try:
            nested_frame = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.TAG_NAME, "iframe"))
            )
            driver.switch_to.frame(nested_frame)
            print("检测到嵌套iframe，已切换")
        except TimeoutException:
            print("未检测到嵌套iframe")

        # 13. 选择BU为DFM
        print("设置BU选项...")
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#select_bu"))
        ).click()
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#select_bu option[value='DFM']"))
        ).click()

        # 14. 选择Category为S
        print("设置Category选项...")
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#select_Category"))
        ).click()
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#select_Category option[value='S']"))
        ).click()

        # 15. 输入产品关键词
        print("输入产品关键词...")
        product_input = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "txt_product"))
        )
        product_input.clear()
        product_input.send_keys("equipment")

        # 16. 点击导出按钮
        print("正在导出数据...")
        export_btn = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@onclick='exportData()']"))
        )
        export_btn.click()

        # 17. 等待下载完成
        print("等待下载完成...")
        max_wait = 60  # 最大等待60秒
        start_time = time.time()
        downloaded = False

        while not downloaded and (time.time() - start_time) < max_wait:
            time.sleep(2)
            files = glob.glob(os.path.join(download_dir, '*Project_List*.*'))
            for f in files:
                if not f.endswith(('.crdownload', '.tmp')):
                    # 检查文件是否完整
                    size1 = os.path.getsize(f)
                    time.sleep(1)
                    size2 = os.path.getsize(f)
                    if size1 == size2 and size1 > 0:
                        print(f"下载完成: {f}")
                        downloaded = True
                        break

        if not downloaded:
            raise TimeoutError("文件下载超时")

    except Exception as e:
        print(f"操作过程中出错: {str(e)}")
        raise
    finally:
        driver.quit()
        print("浏览器已关闭")


def get_latest_project_list():
    """获取最新下载的Project_List文件"""
    # 先执行下载
    download_project_list()

    # 然后查找文件
    download_dir = r'C:\Users\uxuxs004\Downloads'
    pattern = os.path.join(download_dir, '*Project_List*.xls*')
    files = glob.glob(pattern)

    if not files:
        raise FileNotFoundError("未找到Project_List文件")

    # 返回最新文件
    return max(files, key=os.path.getmtime)


def main():
    # 1. 读取Project_List.xls并保留指定列，同时添加筛选条件
    try:
        project_list_path = get_latest_project_list()
        print(f"使用Project_List文件: {project_list_path}")

        df_org = pd.read_excel(
            project_list_path,
            usecols=['Project ID', 'Project Name', 'EBP Leader',
                     'Product Engineer', 'Project State',
                     'Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR',
                     'Phase 4 Gate Exit-CPA', 'Category', 'ORG', 'Product Group'],
            engine='xlrd' if project_list_path.endswith('.xls') else 'openpyxl'
        )
    except Exception as e:
        print(f"读取Project_List文件失败: {str(e)}")
        return

    # 添加筛选条件：Category='S'，ORG='DFM'，Product Group='Equipment'
    df_org = df_org[
        (df_org['Category'] == 'S') &
        (df_org['ORG'] == 'DFM') &
        (df_org['Product Group'] == 'Equipment')
        ]

    # 只保留中文字符（EBP Leader和Product Engineer列）
    df_org['EBP Leader'] = df_org['EBP Leader'].apply(extract_chinese)
    df_org['Product Engineer'] = df_org['Product Engineer'].apply(extract_chinese)

    # 只保留需要的列
    keep_cols = ['Project ID', 'Project Name', 'EBP Leader',
                 'Product Engineer', 'Project State',
                 'Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR',
                 'Phase 4 Gate Exit-CPA']
    df_org = df_org[keep_cols]

    # 2. 读取project.xlsx和project_plan.xlsx
    df_prt = pd.read_excel(r'C:\XSR\githubPage\yanfeng\project.xlsx', engine='openpyxl')
    df_prtp = pd.read_excel(r'C:\XSR\githubPage\yanfeng\project_plan.xlsx', engine='openpyxl')

    # 3. 更新df_prt中与df_prtp匹配的Project ID的数据
    cols_to_update = ['PKR_S', 'PKR', 'P2', 'P3', 'P4', '备注', 'SSO_S', 'SSO_issue']
    for project_id in df_prtp['Project ID']:
        if project_id in df_prt['Project ID'].values:
            mask_prt = df_prt['Project ID'] == project_id
            mask_prtp = df_prtp['Project ID'] == project_id
            for col in cols_to_update:
                if col in df_prtp.columns:
                    df_prt.loc[mask_prt, col] = df_prtp.loc[mask_prtp, col].values[0]

    # 4. 将df_org中不存在于df_prt的Project ID添加到df_prt中
    cols_to_add = ['Project ID', 'Project Name', 'EBP Leader', 'Product Engineer',
                   'Project State', 'Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR',
                   'Phase 4 Gate Exit-CPA']

    # 处理FutureWarning
    new_projects = df_org[~df_org['Project ID'].isin(df_prt['Project ID'])][cols_to_add].dropna(how='all')
    if not new_projects.empty:
        df_prt = pd.concat([df_prt, new_projects], ignore_index=True)

    # 5. 更新df_prt中与df_org匹配的Project ID的数据
    for project_id in df_org['Project ID']:
        if project_id in df_prt['Project ID'].values:
            mask_prt = df_prt['Project ID'] == project_id
            mask_org = df_org['Project ID'] == project_id
            for col in cols_to_add:
                if col in df_org.columns:
                    df_prt.loc[mask_prt, col] = df_org.loc[mask_org, col].values[0]

    # 6. 自动标记项目状态（G/Y/R）
    today = pd.Timestamp(datetime.now().date())  # 统一使用Timestamp类型

    # 处理P2列
    mask_p2 = df_prt['P2'] != 'B'
    df_prt.loc[mask_p2, 'Phase 2 Gate Exit-DVR'] = pd.to_datetime(df_prt.loc[mask_p2, 'Phase 2 Gate Exit-DVR'],
                                                                  errors='coerce')

    for idx, row in df_prt[mask_p2].iterrows():
        if pd.isna(row['Phase 2 Gate Exit-DVR']):
            continue
        delta = (row['Phase 2 Gate Exit-DVR'] - today).days
        if delta > 7:
            df_prt.at[idx, 'P2'] = 'G'
        elif delta >= 0:
            df_prt.at[idx, 'P2'] = 'Y'
        else:
            df_prt.at[idx, 'P2'] = 'R'

    # 处理P3列
    mask_p3 = df_prt['P3'] != 'B'
    df_prt.loc[mask_p3, 'Phase 3 Gate Exit-FPR'] = pd.to_datetime(df_prt.loc[mask_p3, 'Phase 3 Gate Exit-FPR'],
                                                                  errors='coerce')

    for idx, row in df_prt[mask_p3].iterrows():
        if pd.isna(row['Phase 3 Gate Exit-FPR']):
            continue
        delta = (row['Phase 3 Gate Exit-FPR'] - today).days
        if delta > 7:
            df_prt.at[idx, 'P3'] = 'G'
        elif delta >= 0:
            df_prt.at[idx, 'P3'] = 'Y'
        else:
            df_prt.at[idx, 'P3'] = 'R'

    # 处理P4列
    mask_p4 = df_prt['P4'] != 'B'
    df_prt.loc[mask_p4, 'Phase 4 Gate Exit-CPA'] = pd.to_datetime(df_prt.loc[mask_p4, 'Phase 4 Gate Exit-CPA'],
                                                                  errors='coerce')

    for idx, row in df_prt[mask_p4].iterrows():
        if pd.isna(row['Phase 4 Gate Exit-CPA']):
            continue
        delta = (row['Phase 4 Gate Exit-CPA'] - today).days
        if delta > 7:
            df_prt.at[idx, 'P4'] = 'G'
        elif delta >= 0:
            df_prt.at[idx, 'P4'] = 'Y'
        else:
            df_prt.at[idx, 'P4'] = 'R'

    # 7. 筛选当前一周内的数据
    date_cols = ['Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR', 'Phase 4 Gate Exit-CPA']
    for col in date_cols:
        df_prt[col] = pd.to_datetime(df_prt[col], errors='coerce')

    start_of_week = today - pd.Timedelta(days=today.dayofweek)
    end_of_week = start_of_week + pd.Timedelta(days=6)

    mask = ((df_prt[date_cols[0]] >= start_of_week) & (df_prt[date_cols[0]] <= end_of_week)) | \
           ((df_prt[date_cols[1]] >= start_of_week) & (df_prt[date_cols[1]] <= end_of_week)) | \
           ((df_prt[date_cols[2]] >= start_of_week) & (df_prt[date_cols[2]] <= end_of_week))

    df_prtp1 = df_prt[mask].copy()

    # 7.1 新增：从Project Name中提取工作令号（数字前后都是非数字）
    df_prt['工作令号'] = df_prt['Project Name'].apply(extract_six_digits)
    df_prtp1['工作令号'] = df_prtp1['Project Name'].apply(extract_six_digits)

    # 8. 新增：同步数据到飞书项目信息表V5.xlsx
    print("\n开始同步数据到飞书项目信息表...")
    sync_with_feishu_project_table(df_prt)

    # 9. 保存df_prt到project.xlsx
    def save_excel(df, filepath):
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']

                # 定义列宽配置字典 - 添加新列的宽度配置
                column_widths = {
                    'Project ID': 6,
                    'Project Name': 64,
                    'EBP Leader': 6.25,
                    'Product Engineer': 6.25,
                    'Project State': 7.88,
                    'PKR_S': 2.5,
                    'PKR': 2.5,
                    'P2': 2.5,
                    'P3': 2.5,
                    'P4': 2.5,
                    '备注': 2.5,
                    'SSO_S': 2.5,
                    'SSO_issue': 75,
                    'Phase 2 Gate Exit-DVR': 11,
                    'Phase 3 Gate Exit-FPR': 11,
                    'Phase 4 Gate Exit-CPA': 11,
                    '工作令号': 8  # 新增列的宽度
                }

                # 设置各列宽度
                for col_name, width in column_widths.items():
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name)
                        col_letter = get_column_letter(col_idx + 1)
                        worksheet.column_dimensions[col_letter].width = width

                # 设置日期格式（针对日期列）
                date_cols = ['Phase 2 Gate Exit-DVR', 'Phase 3 Gate Exit-FPR', 'Phase 4 Gate Exit-CPA']
                for col in date_cols:
                    if col in df.columns:
                        col_idx = df.columns.get_loc(col)
                        col_letter = get_column_letter(col_idx + 1)
                        for cell in worksheet[col_letter][1:]:  # 跳过标题行
                            cell.number_format = 'YYYY-MM-DD'

            return True
        except Exception as e:
            print(f"保存 {filepath} 时出错: {str(e)}")
            return False

    # 保存文件
    project_path = r'C:\XSR\githubPage\yanfeng\project.xlsx'
    project_plan_path = r'C:\XSR\githubPage\yanfeng\project_plan.xlsx'

    if save_excel(df_prt, project_path):
        print("project.xlsx 保存成功")
    else:
        print("project.xlsx 保存失败")

    if save_excel(df_prtp1, project_plan_path):
        print("project_plan.xlsx 保存成功")
    else:
        print("project_plan.xlsx 保存失败")

    # 10. 自动打开project_plan.xlsx
    try:
        subprocess.Popen([r'C:\XSR\githubPage\yanfeng\project_plan.xlsx'], shell=True)
        print("已打开 project_plan.xlsx")
    except Exception as e:
        print(f"打开文件失败: {str(e)}")


if __name__ == "__main__":
    main()